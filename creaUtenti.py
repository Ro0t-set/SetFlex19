#!/usr/bin/env python
import os
import sys

if __name__ == "__main__":
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "django_project.settings")

    from django.core.management import execute_from_command_line

    execute_from_command_line(sys.argv)

import math
import string
from django.contrib.auth.models import User


#########rddfhgjkl#############

from random import *
import openpyxl
from openpyxl import Workbook
characters =  string.digits
n_effettivo=0
riga = 1417
pw = openpyxl.Workbook()
coordinate = pw.get_sheet_by_name('Sheet')
while riga < 1443:
        riga = riga + 1
        n_effettivo=int(n_effettivo)
        n_effettivo= n_effettivo+1
        password =  "".join(choice(characters) for x in range(randint(8, 8)))

        excel_document = openpyxl.load_workbook('listastudenti.xlsx')
        sheet = excel_document.get_sheet_by_name('Foglio1')
        nome_utente_cognome= str(sheet.cell(row = riga, column = 2).value)
        nome_utente_nome= str(sheet.cell(row = riga, column = 3).value)
        nome_utente= (nome_utente_nome+nome_utente_cognome)
        classe= str(sheet.cell(row = riga, column = 4).value)
        utente=str(nome_utente+classe)


        print (riga)
        print (utente)
        print (password)
        user=User.objects.create_user(utente, password=password)
        user.is_superuser=False
        user.is_staff=False
        user.save()

        posizione_nome = 'A{0}'.format(n_effettivo)
        coordinate[posizione_nome].value = nome_utente
        posizione_pass = 'B{0}'.format(n_effettivo)
        coordinate[posizione_pass].value = password



pw.save('Pw.xlsx')
